#!/usr/bin/env python3
"""
converter_scada.py — Conversão em lote de XLSX (1 sensor por arquivo) para CSV/Parquet.

Layout esperado de cada .xlsx:
    Coluna A : data e hora  (dd/mm/aaaa hh:mm:ss, texto ou datetime nativo do Excel)
    Coluna B : valor        (cabeçalho = TAG do sensor)
    Aba única, sem coluna de qualidade.
    Linhas de metadado acima do cabeçalho são detectadas e descartadas.

Saída:
    <saida>/<TAG>.parquet  -> timestamp datetime64[ns], valor float64 (padrão)
    <saida>/<TAG>.csv      -> timestamp em ISO 8601 (parsing não ambíguo)
    <saida>/_manifest.csv  -> índice de tags para o player (tag, arquivo, n, t_ini, t_fim)

Uso:
    python converter_scada.py --entrada ./xlsx --saida ./dados
    python converter_scada.py --entrada ./xlsx --saida ./dados --formato ambos --jobs 4

Dependências: pandas, openpyxl, pyarrow (para Parquet)
"""

from __future__ import annotations

import argparse
import ctypes
import os
import re
import sys
import unicodedata
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

FORMATO_TS = "%d/%m/%Y %H:%M:%S"
MAX_LINHAS_METADADO = 20  # até onde procurar o cabeçalho

# Pico de RSS medido em ler_planilha ~= 2,5x o XML descomprimido da aba.
# 3,0 dá margem para fragmentação e variação de layout.
FATOR_PICO_MEM = 3.0
FRACAO_RAM_UTIL = 0.6   # nunca comprometer mais que isso da RAM livre
PISO_PICO_BYTES = 64 * 1024**2   # custo mínimo assumido por worker (interpretador + pandas)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def sanitizar(nome: str) -> str:
    """Converte um TAG em nome de arquivo seguro, preservando legibilidade."""
    nfd = unicodedata.normalize("NFD", str(nome))
    ascii_ = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    limpo = re.sub(r"[^A-Za-z0-9._-]+", "_", ascii_).strip("_")
    return limpo or "SEM_TAG"


def parse_timestamp(valor):
    """Aceita datetime nativo do Excel ou string dd/mm/aaaa hh:mm:ss."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    try:
        return datetime.strptime(str(valor).strip(), FORMATO_TS)
    except ValueError:
        # tolera segundos ausentes ou separador com 'T'
        try:
            return pd.to_datetime(str(valor).strip(), dayfirst=True).to_pydatetime()
        except Exception:
            return None


def parse_valor(valor):
    """Aceita float nativo ou texto com vírgula decimal (pt-BR)."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return float(valor)
    txt = str(valor).strip()
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$", txt):      # 1.234.567,89
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Leitura
# --------------------------------------------------------------------------- #

def ler_planilha(caminho: Path) -> tuple[str, pd.DataFrame]:
    """
    Lê o xlsx em modo streaming (read_only) e devolve (tag, DataFrame).
    read_only evita carregar a planilha inteira na memória — importante para
    arquivos grandes de captura por exceção.
    """
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        tag = None
        registros: list[tuple[datetime, float]] = []

        for i, linha in enumerate(ws.iter_rows(values_only=True)):
            if not linha or all(c is None for c in linha[:2]):
                continue

            a = linha[0]
            b = linha[1] if len(linha) > 1 else None

            if tag is None:
                ts = parse_timestamp(a)
                if ts is not None:
                    # chegou nos dados sem header textual: usa o nome do arquivo
                    tag = caminho.stem
                    v = parse_valor(b)
                    if v is not None:
                        registros.append((ts, v))
                    continue
                # candidato a cabeçalho: coluna B com texto não numérico
                if i < MAX_LINHAS_METADADO and b is not None and str(b).strip():
                    if parse_valor(b) is None:
                        tag = str(b).strip()
                continue

            ts = parse_timestamp(a)
            v = parse_valor(b)
            if ts is not None and v is not None:
                registros.append((ts, v))
    finally:
        wb.close()

    if tag is None:
        tag = caminho.stem

    df = pd.DataFrame(registros, columns=["timestamp", "valor"])
    if df.empty:
        return tag, df

    df["timestamp"] = pd.to_datetime(df["timestamp"])
    df["valor"] = df["valor"].astype("float64")

    # ordena e remove timestamps duplicados (mantém o último — semântica de RBE)
    df = df.sort_values("timestamp", kind="mergesort")
    df = df.drop_duplicates(subset="timestamp", keep="last").reset_index(drop=True)

    return tag, df


# --------------------------------------------------------------------------- #
# Escrita
# --------------------------------------------------------------------------- #

def escrever(df: pd.DataFrame, tag: str, saida: Path, formato: str) -> None:
    base = saida / sanitizar(tag)
    if formato in ("csv", "ambos"):
        df.to_csv(
            base.with_suffix(".csv"),
            index=False,
            date_format="%Y-%m-%dT%H:%M:%S",
        )
    if formato in ("parquet", "ambos"):
        df.to_parquet(base.with_suffix(".parquet"), index=False, compression="zstd")


# --------------------------------------------------------------------------- #
# Unidade de trabalho (executada em processo separado)
# --------------------------------------------------------------------------- #

def processar(caminho: Path, saida: Path, formato: str) -> dict:
    """
    Converte um único arquivo. Precisa ser função de nível de módulo para ser
    picklable pelo ProcessPoolExecutor. Devolve a entrada do manifesto.
    """
    tag, df = ler_planilha(caminho)
    if df.empty:
        return {"status": "vazio", "origem": caminho.name}

    escrever(df, tag, saida, formato)
    return {
        "status": "ok",
        "tag": tag,
        "arquivo": sanitizar(tag),
        "origem": caminho.name,
        "n_pontos": len(df),
        "t_inicio": df["timestamp"].iloc[0].isoformat(),
        "t_fim": df["timestamp"].iloc[-1].isoformat(),
        "valor_min": float(df["valor"].min()),
        "valor_max": float(df["valor"].max()),
    }


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def memoria_disponivel_bytes() -> int:
    """RAM livre, de forma portátil. Cai para um valor conservador se não descobrir."""
    try:
        import psutil  # opcional
        return int(psutil.virtual_memory().available)
    except Exception:
        pass
    if sys.platform == "win32":
        try:
            class _MS(ctypes.Structure):
                _fields_ = [("dwLength", ctypes.c_ulong),
                            ("dwMemoryLoad", ctypes.c_ulong),
                            ("ullTotalPhys", ctypes.c_ulonglong),
                            ("ullAvailPhys", ctypes.c_ulonglong),
                            ("ullTotalPageFile", ctypes.c_ulonglong),
                            ("ullAvailPageFile", ctypes.c_ulonglong),
                            ("ullTotalVirtual", ctypes.c_ulonglong),
                            ("ullAvailVirtual", ctypes.c_ulonglong),
                            ("ullAvailExtendedVirtual", ctypes.c_ulonglong)]
            st = _MS()
            st.dwLength = ctypes.sizeof(_MS)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(st)):
                return int(st.ullAvailPhys)
        except Exception:
            pass
    else:
        try:  # Linux/macOS
            return os.sysconf("SC_AVPHYS_PAGES") * os.sysconf("SC_PAGE_SIZE")
        except Exception:
            pass
    return 2 * 1024**3  # fallback conservador: 2 GB


def estimar_pico_bytes(caminho: Path) -> int:
    """
    Estima o pico de memória para converter este arquivo, lendo apenas o índice
    do zip — o tamanho descomprimido da aba está no cabeçalho, sem custo de I/O.
    """
    try:
        with zipfile.ZipFile(caminho) as z:
            abas = [i.file_size for i in z.infolist()
                    if "sheet" in i.filename.lower() and i.filename.endswith(".xml")]
        xml = max(abas) if abas else caminho.stat().st_size * 8
    except Exception:
        xml = caminho.stat().st_size * 8  # zip ilegível: assume 8x o comprimido
    return max(PISO_PICO_BYTES, int(xml * FATOR_PICO_MEM))


def dimensionar_jobs(arquivos: list[Path], verbose: bool = True) -> int:
    """
    Escolhe o número de processos: limitado por cores, pela RAM livre e pelo
    número de arquivos. Usa o maior arquivo como custo por worker (pior caso,
    já que a ordem de agendamento não é garantida).
    """
    picos = [estimar_pico_bytes(c) for c in arquivos]
    pior = max(picos)
    livre = memoria_disponivel_bytes()
    orcamento = int(livre * FRACAO_RAM_UTIL)

    por_cpu = max(1, (os.cpu_count() or 2) // 2)
    por_ram = max(1, orcamento // pior)
    jobs = max(1, min(por_cpu, por_ram, len(arquivos)))

    if verbose:
        print(f"RAM livre {livre/1024**3:.1f} GB | maior arquivo estimado em "
              f"{pior/1024**2:.0f} MB por worker")
        limite = "RAM" if por_ram < por_cpu else "CPU"
        print(f"limites: CPU={por_cpu}, RAM={por_ram}, arquivos={len(arquivos)} "
              f"-> {jobs} processo(s) (gargalo: {limite})")
    return jobs


def main() -> int:
    p = argparse.ArgumentParser(
        description="XLSX SCADA -> CSV/Parquet, um arquivo por sensor"
    )
    p.add_argument("--entrada", type=Path, required=True, help="pasta com os .xlsx")
    p.add_argument("--saida", type=Path, required=True, help="pasta de destino")
    p.add_argument("--formato", choices=["csv", "parquet", "ambos"], default="parquet")
    p.add_argument("--jobs", default="auto",
                   help="processos paralelos: 'auto' (padrão, dimensiona por CPU e "
                        "RAM livre), ou um inteiro; 1 = sequencial")
    args = p.parse_args()

    if not args.entrada.is_dir():
        print(f"Pasta de entrada não encontrada: {args.entrada}", file=sys.stderr)
        return 1
    args.saida.mkdir(parents=True, exist_ok=True)

    arquivos = [c for c in args.entrada.glob("*.xlsx")
                if not c.name.startswith("~$")]  # ignora temporários do Excel
    if not arquivos:
        print(f"Nenhum .xlsx em {args.entrada}", file=sys.stderr)
        return 1

    # maiores primeiro: evita que um arquivo pesado sobre no fim e ocupe um
    # único worker enquanto os outros já terminaram
    arquivos.sort(key=lambda c: c.stat().st_size, reverse=True)

    print(f"{len(arquivos)} arquivo(s), formato '{args.formato}'")
    if str(args.jobs).lower() == "auto":
        jobs = dimensionar_jobs(arquivos)
    else:
        try:
            jobs = max(1, min(int(args.jobs), len(arquivos)))
        except ValueError:
            print(f"--jobs inválido: {args.jobs!r} (use 'auto' ou um inteiro)",
                  file=sys.stderr)
            return 1
        print(f"{jobs} processo(s) (definido manualmente)")
    print()

    manifesto: list[dict] = []
    erros = 0

    if jobs == 1:
        for caminho in arquivos:
            try:
                r = processar(caminho, args.saida, args.formato)
            except Exception as exc:
                print(f"[ERRO ] {caminho.name}: {exc}", file=sys.stderr)
                erros += 1
                continue
            _registrar(r, manifesto)
    else:
        with ProcessPoolExecutor(max_workers=jobs) as pool:
            futuros = {
                pool.submit(processar, c, args.saida, args.formato): c
                for c in arquivos
            }
            for fut in as_completed(futuros):
                caminho = futuros[fut]
                try:
                    r = fut.result()
                except Exception as exc:
                    print(f"[ERRO ] {caminho.name}: {exc}", file=sys.stderr)
                    erros += 1
                    continue
                _registrar(r, manifesto)

    if manifesto:
        # as_completed devolve fora de ordem — reordena para o manifesto ficar estável
        df_man = pd.DataFrame(manifesto).sort_values("tag").reset_index(drop=True)
        destino = args.saida / "_manifest.csv"
        df_man.to_csv(destino, index=False)
        print(f"\n{len(df_man)} sensor(es) convertido(s). Manifesto: {destino}")

    if erros:
        print(f"{erros} arquivo(s) com erro.", file=sys.stderr)
    return 1 if erros and not manifesto else 0


def _registrar(r: dict, manifesto: list[dict]) -> None:
    if r["status"] == "vazio":
        print(f"[VAZIO] {r['origem']}: nenhuma linha válida")
        return
    manifesto.append({k: v for k, v in r.items() if k != "status"})
    print(f"[OK   ] {r['origem']} -> {r['arquivo']} ({r['n_pontos']:,} pontos)")


if __name__ == "__main__":
    raise SystemExit(main())
